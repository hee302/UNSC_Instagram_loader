import openpyxl
import instaloader
import time
import random
import os
import sys

def resource_path(relative_path):
    """ exe 실행 시 임시 폴더 경로 해결을 위한 함수 """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

print("=" * 40)
print("   인스타그램 다 들어가보기 귀찮은 운영진을 위한 인스타그램 게시글 수 추출기")
print("=" * 40)

# 1. 사용자로부터 설정값 입력받기
target_file = input("1. 엑셀 파일명을 입력하세요 (예: instagram.xlsx): ").strip()

# 파일 존재 여부 확인
if not os.path.exists(target_file):
    print(f"\n[오류] '{target_file}' 파일을 찾을 수 없습니다.")
    print("exe 파일과 같은 폴더에 엑셀 파일이 있는지 확인해주세요.")
    input("엔터를 누르면 종료합니다...")
    exit()

read_col = input("2. 아이디가 있는 [열] 알파벳은? (예: C): ").strip().upper()
write_col = input("3. 결과를 저장할 [열] 알파벳은? (예: G): ").strip().upper()

try:
    start_row = int(input("4. 시작할 [행] 번호는? (예: 5): ").strip())
    end_row = int(input("5. 끝낼 [행] 번호는? (예: 32): ").strip())
except ValueError:
    print("\n[오류] 행 번호는 숫자만 입력해야 합니다.")
    input("엔터를 누르면 종료합니다...")
    exit()

print("\n" + "=" * 40)
print(f"설정 확인: {read_col}{start_row} ~ {read_col}{end_row} 읽기 -> {write_col}열에 쓰기")
print("로그인을 하지 않으므로 천천히 진행됩니다. (잠시 대기...)")
print("=" * 40)
time.sleep(2)

# 2. 엑셀 로드 및 로더 초기화
wb = openpyxl.load_workbook(target_file)
ws = wb.active
L = instaloader.Instaloader()

error_count = 0

# 3. 반복 작업 시작
for row in range(start_row, end_row + 1):
    # 사용자가 입력한 열(read_col)과 행(row)을 조합 (예: C + 5 -> C5)
    cell_id = ws[f'{read_col}{row}'].value
    
    if cell_id is None:
        print(f"[행 {row}] 빈 칸입니다. 건너뜁니다.")
        continue

    username = str(cell_id).strip()

    try:
        profile = instaloader.Profile.from_username(L.context, username)
        count = profile.mediacount
        
        # 사용자가 입력한 결과 열(write_col)에 저장
        ws[f'{write_col}{row}'].value = count
        print(f"[행 {row}] {username} : {count}개 완료")
        
        error_count = 0 # 성공 시 에러 카운트 초기화

        # 중간 저장 (파일명 앞에 'result_'를 붙여서 저장)
        if row % 5 == 0:
            wb.save(f'result_{target_file}')
            print("   >> 중간 저장 완료")

    except Exception as e:
        print(f"[행 {row}] {username} : 실패 ({e})")
        ws[f'{write_col}{row}'].value = "Error"
        error_count += 1
        
        if error_count >= 3:
            print("\n!!! 연속 3회 에러. 차단 방지를 위해 여기서 멈춥니다. !!!")
            break

    # 랜덤 대기 (15~30초)
    wait_time = random.uniform(15, 30)
    print(f"   ({int(wait_time)}초 대기 중...)")
    time.sleep(wait_time)

# 4. 최종 저장
output_filename = f'result_{target_file}'
wb.save(output_filename)

print("\n" + "=" * 40)
print(f"완료되었습니다! '{output_filename}' 파일을 확인하세요.")
print("=" * 40)
input("종료하려면 엔터 키를 누르세요...")